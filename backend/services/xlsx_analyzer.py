from openpyxl import load_workbook


class XLSXAnalyzer:

    def extract(self, file_path: str):

        workbook = load_workbook(
            file_path,
            read_only=True
        )

        content = []

        for sheet in workbook.sheetnames:

            worksheet = workbook[sheet]

            for row in worksheet.iter_rows(
                values_only=True
            ):

                values = [
                    str(cell)
                    for cell in row
                    if cell is not None
                ]

                if values:

                    content.append(
                        " ".join(values)
                    )

        return "\n".join(content)